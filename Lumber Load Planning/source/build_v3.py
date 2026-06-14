from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule, CellIsRule
from openpyxl.utils import get_column_letter
import importlib.util
spec=importlib.util.spec_from_file_location("sv3","/home/claude/solver_v3.py"); sv3=importlib.util.module_from_spec(spec); spec.loader.exec_module(sv3)
PRODUCTS=sv3.PRODUCTS; LENGTHS=sv3.LENGTHS; GRADES=sv3.GRADES; SEP="-"
PCOLORS={"2x4":"3A7CB8","2x6":"3D8A4F","2x8":"C8892A","2x10":"7A4FA0","4x4":"B0473C","4x6":"2E8A8A","6x6":"6B4A2F"}
GCOLORS={"1":"1A7A1A","2":"2A5AAA","3":"D08020","4":"C03020","2P":"7A3AA0","MSR":"0A8A8A"}
# Color hierarchy: LENGTH = primary (fill), PRODUCT = secondary (thick border), GRADE = tertiary (font color)
LCOLORS={8:"3A6E9E",10:"2E7A5E",12:"7A5E2E",14:"6E3A7A",16:"8A3A3A",18:"2E7A6E",20:"4A6E2E"}
# light, high-contrast grade tints readable as font color on the dark length fills
GFONT={"1":"BFF2BF","2":"BDD7FF","3":"FFD79E","4":"FFC0B8","2P":"E6C2FA","MSR":"A8EEEE"}
DEFAULT=sv3.DEFAULT_ITEMS; DEFAULT_ROWS=10
SOL,_=sv3.solve(DEFAULT,DEFAULT_ROWS)
MAX_ROWS=14; MAX_SLOTS=9; LI_COUNT=20

wb=Workbook()
def F(**k): return Font(name='Arial',**k)
hdr=PatternFill('solid',fgColor='222820'); sub=PatternFill('solid',fgColor='2B3126'); yellow=PatternFill('solid',fgColor='FFFF99')
thin=Side(style='thin',color='B0B0B0'); border=Border(left=thin,right=thin,top=thin,bottom=thin)
center=Alignment(horizontal='center',vertical='center'); left=Alignment(horizontal='left',vertical='center')
def gborder(hexc): s=Side(style='medium',color=hexc); return Border(left=s,right=s,top=s,bottom=s)

ws=wb.active; ws.title="Planner"; ws.sheet_view.showGridLines=False
ws['A1']="CENTERBEAM LUMBER CAR LAYOUT PLANNER  —  72 ft"; ws['A1'].font=F(size=15,bold=True,color='2E5E20'); ws.merge_cells('A1:N1'); ws['A1'].alignment=left
ws.row_dimensions[1].height=24
ws['A2']="Inventory is a line-item tally: each line = Product + Length + Grade + Packs. Cell FILL = product, BORDER = grade. Each car row totals exactly 72 ft. Example loaded."
ws['A2'].font=F(size=9,italic=True,color='666666'); ws.merge_cells('A2:N2')

# Config
ws['A4']="CONFIGURATION"; ws['A4'].font=F(size=10,bold=True,color='FFFFFF'); ws['A4'].fill=hdr; ws.merge_cells('A4:C4')
ws['A5']="Number of Rows:"; ws['A5'].font=F(size=11,bold=True)
ws['C5']=DEFAULT_ROWS; ws['C5'].font=F(size=11,bold=True,color='0000FF'); ws['C5'].fill=yellow; ws['C5'].border=border; ws['C5'].alignment=center
dvr=DataValidation(type="list",formula1='"10,14"',allow_blank=False); ws.add_data_validation(dvr); dvr.add(ws['C5'])
ws['A6']="Row Capacity (ft):"; ws['A6'].font=F(size=11); ws['C6']=72; ws['C6'].font=F(size=11); ws['C6'].border=border; ws['C6'].alignment=center

# ---- LINE ITEM INVENTORY ----
ws['A8']="LUMBER INVENTORY  —  line-item tally (Product × Length × Grade)"; ws['A8'].font=F(size=10,bold=True,color='FFFFFF'); ws['A8'].fill=hdr; ws.merge_cells('A8:H8')
LI_HDR=9
heads=["#","Product","Length","Grade","Packs","Lineal","Placed","Remain"]
widths=[4,9,8,7,7,8,8,8]
for j,(h,w) in enumerate(zip(heads,widths)):
    c=ws.cell(LI_HDR,1+j,h); c.font=F(size=9,bold=True,color='FFFFFF'); c.fill=sub; c.alignment=center; c.border=border
    ws.column_dimensions[get_column_letter(1+j)].width=w
LI_FIRST=LI_HDR+1; LI_LAST=LI_FIRST+LI_COUNT-1
GF=LI_LAST+5  # set grid first later precisely; placeholder
# We'll know grid after; build grid range string now using planned GF
GF=LI_LAST+5
GL=GF+MAX_ROWS-1
GRID=f"$B${GF}:$J${GL}"
CHKR=f"$L${GF}:$L${GL}"

dvP=DataValidation(type="list",formula1='"%s"'%",".join(PRODUCTS),allow_blank=True); ws.add_data_validation(dvP)
dvL=DataValidation(type="list",formula1='"%s"'%",".join(str(x) for x in LENGTHS),allow_blank=True); ws.add_data_validation(dvL)
dvG=DataValidation(type="list",formula1='"%s"'%",".join(GRADES),allow_blank=True); ws.add_data_validation(dvG)

for i in range(LI_COUNT):
    r=LI_FIRST+i
    ws.cell(r,1,i+1).font=F(size=8,color='999999'); ws.cell(r,1).alignment=center; ws.cell(r,1).border=border
    for col in range(2,5):
        cell=ws.cell(r,col); cell.alignment=center; cell.border=border; cell.fill=yellow; cell.font=F(color='0000FF')
    pk=ws.cell(r,5); pk.alignment=center; pk.border=border; pk.fill=yellow; pk.font=F(color='0000FF')
    ws.cell(r,6,f'=IF(B{r}="","",C{r}*E{r})').alignment=center; ws.cell(r,6).border=border
    ws.cell(r,7,f'=IF(B{r}="","",COUNTIF({GRID},B{r}&"{SEP}"&C{r}&"{SEP}"&D{r}))').alignment=center; ws.cell(r,7).border=border
    ws.cell(r,8,f'=IF(B{r}="","",E{r}-G{r})').alignment=center; ws.cell(r,8).border=border
    ws.cell(r,9,f'=IF(B{r}="",0,C{r}*G{r})').alignment=center  # numeric placed-lineal helper (for summaries)
    dvP.add(ws.cell(r,2)); dvL.add(ws.cell(r,3)); dvG.add(ws.cell(r,4))
# default items
for i,it in enumerate(DEFAULT):
    r=LI_FIRST+i
    ws.cell(r,2,it['product']); ws.cell(r,3,it['length']); ws.cell(r,4,it['grade']); ws.cell(r,5,it['qty'])
LI_TOT=LI_LAST+1
ws.cell(LI_TOT,1,"TOTAL").font=F(bold=True); ws.cell(LI_TOT,1).alignment=center; ws.cell(LI_TOT,1).border=border
ws.merge_cells(start_row=LI_TOT,start_column=1,end_row=LI_TOT,end_column=4)
for col,letter in [(5,'E'),(6,'F'),(7,'G'),(8,'H')]:
    ws.cell(LI_TOT,col,f"=SUM({letter}{LI_FIRST}:{letter}{LI_LAST})").font=F(bold=True); ws.cell(LI_TOT,col).alignment=center; ws.cell(LI_TOT,col).border=border
ws.cell(LI_TOT,9,f"=SUM(I{LI_FIRST}:I{LI_LAST})")  # helper total
ws.column_dimensions['I'].width=2  # keep helper unobtrusive

# ---- STATUS box (cols J..M) ----
SB=10
ws.cell(LI_HDR,SB,"LOAD STATUS").font=F(size=10,bold=True,color='FFFFFF'); ws.cell(LI_HDR,SB).fill=hdr
ws.merge_cells(start_row=LI_HDR,start_column=SB,end_row=LI_HDR,end_column=SB+3)
status=[("Total Lineal Ft:",f"=F{LI_TOT}"),("Car Capacity (ft):","=C5*C6"),
        ("Difference:",f"=F{LI_TOT}-C5*C6"),("Rows Needed:",f"=IF(F{LI_TOT}=0,0,ROUNDUP(F{LI_TOT}/C6,0))"),
        ("Rows OK (=72):",f'=COUNTIF({CHKR},"OK 72")'),("Packs Placed:",f"=COUNTA({GRID})"),
        ("Packs Unplaced:",f"=E{LI_TOT}-COUNTA({GRID})")]
for i,(lbl,frm) in enumerate(status):
    r=LI_HDR+1+i
    ws.cell(r,SB,lbl).font=F(size=10,bold=True); ws.merge_cells(start_row=r,start_column=SB,end_row=r,end_column=SB+1)
    cc=ws.cell(r,SB+2,frm); cc.font=F(size=10,bold=True); cc.alignment=center; cc.border=border
    ws.merge_cells(start_row=r,start_column=SB+2,end_row=r,end_column=SB+3)
VR=LI_HDR+1+len(status)
ws.cell(VR,SB,"STATUS:").font=F(size=11,bold=True); ws.merge_cells(start_row=VR,start_column=SB,end_row=VR,end_column=SB+1)
verdict=(f'=IF(E{LI_TOT}-COUNTA({GRID})>0,"OVERLOADED — "&(E{LI_TOT}-COUNTA({GRID}))&" unplaced",'
         f'IF(AND(COUNTIF({CHKR},"OK 72")=C5,C5>0),"PERFECT — all "&C5&" rows = 72ft",'
         f'IF(COUNTIF({CHKR},"OVER*")>0,"ROW OVER 72 — fix","INCOMPLETE — keep filling")))')
vc=ws.cell(VR,SB+2,verdict); vc.font=F(size=11,bold=True); vc.alignment=center; vc.border=border
ws.merge_cells(start_row=VR,start_column=SB+2,end_row=VR,end_column=SB+3)
for col in ['J','K','L','M']: ws.column_dimensions[col].width=10

# ---- GRID ----
GT=GF-2
ws.cell(GT,1,"LAYOUT GRID  —  product"+SEP+"length"+SEP+"grade per slot;  each row = exactly 72 ft").font=F(size=10,bold=True,color='FFFFFF')
ws.cell(GT,1).fill=hdr; ws.merge_cells(start_row=GT,start_column=1,end_row=GT,end_column=12)
GHR=GF-1
ws.cell(GHR,1,"Row").font=F(size=9,bold=True,color='FFFFFF'); ws.cell(GHR,1).fill=sub; ws.cell(GHR,1).alignment=center; ws.cell(GHR,1).border=border
for s in range(MAX_SLOTS):
    c=ws.cell(GHR,2+s,f"P{s+1}"); c.font=F(size=9,bold=True,color='FFFFFF'); c.fill=sub; c.alignment=center; c.border=border
ws.cell(GHR,11,"Total").font=F(size=9,bold=True,color='FFFFFF'); ws.cell(GHR,11).fill=sub; ws.cell(GHR,11).alignment=center; ws.cell(GHR,11).border=border
ws.cell(GHR,12,"Check").font=F(size=9,bold=True,color='FFFFFF'); ws.cell(GHR,12).fill=sub; ws.cell(GHR,12).alignment=center; ws.cell(GHR,12).border=border
def lenf(c): return f'IFERROR(VALUE(MID({c},FIND("{SEP}",{c})+1,FIND("{SEP}",{c},FIND("{SEP}",{c})+1)-FIND("{SEP}",{c})-1)),0)'
for i in range(MAX_ROWS):
    r=GF+i
    rl=ws.cell(r,1,f"R{i+1}"); rl.font=F(bold=True,color='2E5E20'); rl.alignment=center; rl.border=border
    sol=SOL[i] if i<len(SOL) else []
    for s in range(MAX_SLOTS):
        cell=ws.cell(r,2+s)
        if s<len(sol):
            p,L,g=sol[s]; cell.value=f"{p}{SEP}{L}{SEP}{g}"
        cell.alignment=center; cell.border=border; cell.font=F(bold=True,size=8,color='FFFFFF')
    tot=ws.cell(r,11,"="+"+".join(lenf(f"{get_column_letter(2+s)}{r}") for s in range(MAX_SLOTS))); tot.font=F(bold=True); tot.alignment=center; tot.border=border
    chk=ws.cell(r,12,(f'=IF(ROW()-{GF-1}>$C$5,"— unused —",IF(K{r}=0,"empty",IF(K{r}=$C$6,"OK 72",IF(K{r}>$C$6,"OVER "&K{r},"SHORT "&($C$6-K{r})))))'))
    chk.font=F(bold=True); chk.alignment=center; chk.border=border
for s in range(MAX_SLOTS): ws.column_dimensions[get_column_letter(2+s)].width=12
ws.column_dimensions['K'].width=7; ws.column_dimensions['L'].width=12

# ---- CF: LENGTH fill (primary) + PRODUCT border (secondary) + GRADE font (tertiary) ----
def pborder(hexc): s=Side(style='thick',color=hexc); return Border(left=s,right=s,top=s,bottom=s)
def pid(c): return f'IFERROR(LEFT({c},FIND("{SEP}",{c})-1),"")'
def lid(c): return f'IFERROR(MID({c},FIND("{SEP}",{c})+1,FIND("{SEP}",{c},FIND("{SEP}",{c})+1)-FIND("{SEP}",{c})-1),"")'
def gid(c): return f'IFERROR(MID({c},FIND("{SEP}",{c},FIND("{SEP}",{c})+1)+1,9),"")'
TL=f"B{GF}"
# primary: fill by length
for L,hexc in LCOLORS.items():
    ws.conditional_formatting.add(GRID, FormulaRule(formula=[f'AND({TL}<>"",{lid(TL)}="{L}")'],fill=PatternFill('solid',fgColor=hexc)))
# secondary: thick border by product
for p,hexc in PCOLORS.items():
    ws.conditional_formatting.add(GRID, FormulaRule(formula=[f'AND({TL}<>"",{pid(TL)}="{p}")'],border=pborder(hexc)))
# tertiary: font color by grade (light tints, readable on dark length fills)
for g,hexc in GFONT.items():
    ws.conditional_formatting.add(GRID, FormulaRule(formula=[f'AND({TL}<>"",{gid(TL)}="{g}")'],font=Font(name='Arial',bold=True,size=8,color=hexc)))
# status coloring
for cond,fill,fcol in [('LEFT(L%d,2)="OK"'%GF,'D8F0D0','1A6010'),('LEFT(L%d,4)="OVER"'%GF,'F4C0C0','A30000'),('LEFT(L%d,5)="SHORT"'%GF,'FBE6C0','8A5A00')]:
    ws.conditional_formatting.add(CHKR, FormulaRule(formula=[cond],fill=PatternFill('solid',fgColor=fill),font=Font(name='Arial',color=fcol,bold=True)))
ws.conditional_formatting.add(CHKR, FormulaRule(formula=[f'L{GF}="— unused —"'],fill=PatternFill('solid',fgColor='EDEDED'),font=Font(name='Arial',color='999999',italic=True)))
ws.conditional_formatting.add(GRID, FormulaRule(formula=[f'ROW()-{GF-1}>$C$5'],fill=PatternFill('solid',fgColor='F2F2F2')))
# remaining negative/positive on line items
ws.conditional_formatting.add(f"H{LI_FIRST}:H{LI_LAST}", CellIsRule(operator='lessThan',formula=['0'],fill=PatternFill('solid',fgColor='F4C0C0'),font=Font(name='Arial',color='A30000',bold=True)))

# ---- Legends (LENGTH primary fill, PRODUCT secondary border, GRADE tertiary font) ----
LG=GL+2
ws.cell(LG,1,"LENGTH (fill):").font=F(size=9,bold=True,color='666666')
for i,L in enumerate(LENGTHS):
    c=ws.cell(LG,2+i,f"{L}'"); c.font=F(bold=True,size=9,color='FFFFFF'); c.fill=PatternFill('solid',fgColor=LCOLORS[L]); c.alignment=center; c.border=border
LG2=LG+1
ws.cell(LG2,1,"PRODUCT (border):").font=F(size=9,bold=True,color='666666')
for i,p in enumerate(PRODUCTS):
    c=ws.cell(LG2,2+i,p); c.font=F(bold=True,size=9,color='333333'); c.alignment=center; c.border=pborder(PCOLORS[p])
LG2b=LG2+1
ws.cell(LG2b,1,"GRADE (text color):").font=F(size=9,bold=True,color='666666')
for i,g in enumerate(GRADES):
    c=ws.cell(LG2b,2+i,g); c.font=F(bold=True,size=9,color='FFFFFF'); c.fill=PatternFill('solid',fgColor='555555'); c.alignment=center; c.border=border
    c.font=F(bold=True,size=9,color=GFONT[g])
LG3=LG2b+1
ws.cell(LG3,1,"START OVER: clear grid (%s) and line items (B%d:E%d). One-click Solve/Clear: see 'Auto-Solve (VBA)'."%(GRID.replace("$",""),LI_FIRST,LI_LAST)).font=F(size=9,italic=True,color='666666')
ws.merge_cells(start_row=LG3,start_column=1,end_row=LG3,end_column=12)
ws.column_dimensions['A'].width=8

import json
json.dump({'LI_FIRST':LI_FIRST,'LI_LAST':LI_LAST,'LI_TOT':LI_TOT,'GF':GF,'GL':GL,'GRID':GRID,'CHKR':CHKR},open('/home/claude/meta3.json','w'))
wb.save('/home/claude/cb3.xlsx')
print(f"Planner v3 built. LI {LI_FIRST}-{LI_LAST}, TOT {LI_TOT}, GRID {GRID}, legends {LG}-{LG3}")
