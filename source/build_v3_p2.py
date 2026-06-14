import json, importlib.util
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import FormulaRule, CellIsRule
from openpyxl.utils import get_column_letter
spec=importlib.util.spec_from_file_location("sv3","/home/claude/solver_v3.py"); sv3=importlib.util.module_from_spec(spec); spec.loader.exec_module(sv3)
PRODUCTS=sv3.PRODUCTS; LENGTHS=sv3.LENGTHS; GRADES=sv3.GRADES; SEP="-"
PCOLORS={"2x4":"3A7CB8","2x6":"3D8A4F","2x8":"C8892A","2x10":"7A4FA0","4x4":"B0473C","4x6":"2E8A8A","6x6":"6B4A2F"}
GCOLORS={"1":"1A7A1A","2":"2A5AAA","3":"D08020","4":"C03020","2P":"7A3AA0","MSR":"0A8A8A"}
# hierarchy: LENGTH primary (fill), PRODUCT secondary (thick border), GRADE tertiary (font)
LC={8:"3A6E9E",10:"2E7A5E",12:"7A5E2E",14:"6E3A7A",16:"8A3A3A",18:"2E7A6E",20:"4A6E2E"}
GFONT={"1":"BFF2BF","2":"BDD7FF","3":"FFD79E","4":"FFC0B8","2P":"E6C2FA","MSR":"A8EEEE"}
m=json.load(open('/home/claude/meta3.json'))
LI_FIRST,LI_LAST,LI_TOT,GF,GL,GRID,CHKR=m['LI_FIRST'],m['LI_LAST'],m['LI_TOT'],m['GF'],m['GL'],m['GRID'],m['CHKR']
P="Planner"
wb=load_workbook('/home/claude/cb3.xlsx')
def F(**k): return Font(name='Arial',**k)
hdr=PatternFill('solid',fgColor='222820'); sub=PatternFill('solid',fgColor='2B3126')
thin=Side(style='thin',color='B0B0B0'); border=Border(left=thin,right=thin,top=thin,bottom=thin)
center=Alignment(horizontal='center',vertical='center'); left=Alignment(horizontal='left',vertical='center')
def gborder(hexc): s=Side(style='medium',color=hexc); return Border(left=s,right=s,top=s,bottom=s)
def pborder(hexc): s=Side(style='thick',color=hexc); return Border(left=s,right=s,top=s,bottom=s)

# ===== MANIFEST =====
mf=wb.create_sheet("Manifest"); mf.sheet_view.showGridLines=False
mf['A1']="LOAD MANIFEST"; mf['A1'].font=F(size=15,bold=True,color='2E5E20'); mf.merge_cells('A1:F1')
mf['A2']="Car layout mirrors the Planner. Cell FILL = product, BORDER = grade. Differentiated by product, length, and grade below."
mf['A2'].font=F(size=9,italic=True,color='666666'); mf.merge_cells('A2:L2')
mf['A4']="Car Type:"; mf['A4'].font=F(bold=True); mf['B4']="Centerbeam 72 ft"
mf['A5']="Rows:"; mf['A5'].font=F(bold=True); mf['B5']=f"='{P}'!C5"
mf['A6']="Total Lineal Ft:"; mf['A6'].font=F(bold=True); mf['B6']=f"='{P}'!F{LI_TOT}"
mf['A7']="Date:"; mf['A7'].font=F(bold=True); mf['B7']="=TODAY()"; mf['B7'].number_format='yyyy-mm-dd'
mf['A8']="Notes:"; mf['A8'].font=F(bold=True); mf['B8']="(car #, shipper, instructions)"; mf['B8'].font=F(italic=True,color='999999')

# legends (LENGTH primary fill, PRODUCT secondary border, GRADE tertiary font)
lg=10
mf.cell(lg,1,"LENGTH (fill):").font=F(size=9,bold=True,color='666666')
for i,L in enumerate(LENGTHS):
    c=mf.cell(lg,2+i,f"{L}'"); c.font=F(bold=True,size=9,color='FFFFFF'); c.fill=PatternFill('solid',fgColor=LC[L]); c.alignment=center; c.border=border
mf.cell(lg+1,1,"PRODUCT (border):").font=F(size=9,bold=True,color='666666')
for i,p in enumerate(PRODUCTS):
    c=mf.cell(lg+1,2+i,p); c.font=F(bold=True,size=9,color='333333'); c.alignment=center; c.border=pborder(PCOLORS[p])
mf.cell(lg+2,1,"GRADE (text):").font=F(size=9,bold=True,color='666666')
for i,g in enumerate(GRADES):
    c=mf.cell(lg+2,2+i,g); c.fill=PatternFill('solid',fgColor='555555'); c.alignment=center; c.border=border; c.font=F(bold=True,size=9,color=GFONT[g])

# CAR LAYOUT grid
ct=lg+3
mf.cell(ct,1,"CAR LAYOUT (as loaded)  —  each row = 72 ft, loaded left→right from the center beam").font=F(size=10,bold=True,color='FFFFFF')
mf.cell(ct,1).fill=hdr; mf.merge_cells(start_row=ct,start_column=1,end_row=ct,end_column=12)
mhr=ct+1
mf.cell(mhr,1,"Row").font=F(size=9,bold=True,color='FFFFFF'); mf.cell(mhr,1).fill=sub; mf.cell(mhr,1).alignment=center; mf.cell(mhr,1).border=border
for s in range(9):
    c=mf.cell(mhr,2+s,f"P{s+1}"); c.font=F(size=9,bold=True,color='FFFFFF'); c.fill=sub; c.alignment=center; c.border=border
mf.cell(mhr,11,"Total").font=F(size=9,bold=True,color='FFFFFF'); mf.cell(mhr,11).fill=sub; mf.cell(mhr,11).alignment=center; mf.cell(mhr,11).border=border
mf.cell(mhr,12,"Status").font=F(size=9,bold=True,color='FFFFFF'); mf.cell(mhr,12).fill=sub; mf.cell(mhr,12).alignment=center; mf.cell(mhr,12).border=border
mgf=mhr+1
for i in range(14):
    r=mgf+i; gr=GF+i
    mf.cell(r,1,f"R{i+1}").font=F(bold=True,color='2E5E20'); mf.cell(r,1).alignment=center; mf.cell(r,1).border=border
    for s in range(9):
        col=get_column_letter(2+s)
        cell=mf.cell(r,2+s,f'=IF(\'{P}\'!{col}{gr}="","",\'{P}\'!{col}{gr})'); cell.font=F(bold=True,size=8,color='FFFFFF'); cell.alignment=center; cell.border=border
    mf.cell(r,11,f'=IF(\'{P}\'!K{gr}=0,"",\'{P}\'!K{gr})').alignment=center; mf.cell(r,11).border=border; mf.cell(r,11).font=F(bold=True)
    mf.cell(r,12,f"='{P}'!L{gr}").alignment=center; mf.cell(r,12).border=border
mgl=mgf+13
MGRID=f"$B${mgf}:$J${mgl}"; MCHK=f"$L${mgf}:$L${mgl}"
TL=f"B{mgf}"
pid=f'IFERROR(LEFT({TL},FIND("{SEP}",{TL})-1),"")'
lidf=f'IFERROR(MID({TL},FIND("{SEP}",{TL})+1,FIND("{SEP}",{TL},FIND("{SEP}",{TL})+1)-FIND("{SEP}",{TL})-1),"")'
gid=f'IFERROR(MID({TL},FIND("{SEP}",{TL},FIND("{SEP}",{TL})+1)+1,9),"")'
for L,hexc in LC.items():
    mf.conditional_formatting.add(MGRID, FormulaRule(formula=[f'AND({TL}<>"",{lidf}="{L}")'],fill=PatternFill('solid',fgColor=hexc)))
for p,hexc in PCOLORS.items():
    mf.conditional_formatting.add(MGRID, FormulaRule(formula=[f'AND({TL}<>"",{pid}="{p}")'],border=pborder(hexc)))
for g,hexc in GFONT.items():
    mf.conditional_formatting.add(MGRID, FormulaRule(formula=[f'AND({TL}<>"",{gid}="{g}")'],font=Font(name='Arial',bold=True,size=8,color=hexc)))
for cond,fill,fcol in [('LEFT(L%d,2)="OK"'%mgf,'D8F0D0','1A6010'),('LEFT(L%d,4)="OVER"'%mgf,'F4C0C0','A30000'),('LEFT(L%d,5)="SHORT"'%mgf,'FBE6C0','8A5A00')]:
    mf.conditional_formatting.add(MCHK, FormulaRule(formula=[cond],fill=PatternFill('solid',fgColor=fill),font=Font(name='Arial',color=fcol,bold=True)))
mf.conditional_formatting.add(MCHK, FormulaRule(formula=[f'L{mgf}="— unused —"'],fill=PatternFill('solid',fgColor='EDEDED'),font=Font(name='Arial',color='999999',italic=True)))
mf.conditional_formatting.add(MGRID, FormulaRule(formula=[f"ROW()-{mgf-1}>'{P}'!$C$5"],fill=PatternFill('solid',fgColor='F2F2F2')))
mf.column_dimensions['A'].width=8
for s in range(9): mf.column_dimensions[get_column_letter(2+s)].width=12
mf.column_dimensions['K'].width=7; mf.column_dimensions['L'].width=12

# helper: count placed by criteria using SUMPRODUCT on line items (Placed col G already per item)
# Summary by Product
sp=mgl+2
mf.cell(sp,1,"SUMMARY BY PRODUCT").font=F(size=10,bold=True,color='FFFFFF'); mf.cell(sp,1).fill=hdr; mf.merge_cells(start_row=sp,start_column=1,end_row=sp,end_column=3)
sph=sp+1
for j,h in enumerate(["Product","Placed Packs","Lineal Ft"]):
    c=mf.cell(sph,1+j,h); c.font=F(size=9,bold=True,color='FFFFFF'); c.fill=sub; c.alignment=center; c.border=border
for i,p in enumerate(PRODUCTS):
    r=sph+1+i
    c=mf.cell(r,1,p); c.font=F(bold=True,color='FFFFFF'); c.fill=PatternFill('solid',fgColor=PCOLORS[p]); c.alignment=center; c.border=border
    mf.cell(r,2,f'=SUMIF(\'{P}\'!$B${LI_FIRST}:$B${LI_LAST},"{p}",\'{P}\'!$G${LI_FIRST}:$G${LI_LAST})').alignment=center; mf.cell(r,2).border=border
    mf.cell(r,3,f'=SUMIF(\'{P}\'!$B${LI_FIRST}:$B${LI_LAST},"{p}",\'{P}\'!$I${LI_FIRST}:$I${LI_LAST})').alignment=center; mf.cell(r,3).border=border

# Summary by Grade (placed beside Summary by Product, columns E-G)
GC0=5
mf.cell(sp,GC0,"SUMMARY BY GRADE").font=F(size=10,bold=True,color='FFFFFF'); mf.cell(sp,GC0).fill=hdr
mf.merge_cells(start_row=sp,start_column=GC0,end_row=sp,end_column=GC0+2)
for j,h in enumerate(["Grade","Placed Packs","Lineal Ft"]):
    c=mf.cell(sph,GC0+j,h); c.font=F(size=9,bold=True,color='FFFFFF'); c.fill=sub; c.alignment=center; c.border=border
for i,g in enumerate(GRADES):
    r=sph+1+i
    c=mf.cell(r,GC0,g); c.font=F(bold=True,color=GCOLORS[g]); c.alignment=center; c.border=gborder(GCOLORS[g])
    mf.cell(r,GC0+1,f'=SUMIF(\'{P}\'!$D${LI_FIRST}:$D${LI_LAST},"{g}",\'{P}\'!$G${LI_FIRST}:$G${LI_LAST})').alignment=center; mf.cell(r,GC0+1).border=border
    mf.cell(r,GC0+2,f'=SUMIF(\'{P}\'!$D${LI_FIRST}:$D${LI_LAST},"{g}",\'{P}\'!$I${LI_FIRST}:$I${LI_LAST})').alignment=center; mf.cell(r,GC0+2).border=border

# Product x Grade placed matrix (below both summaries)
mx=sph+1+len(PRODUCTS)+1
mf.cell(mx,1,"PLACED PACKS — product × grade").font=F(size=10,bold=True,color='FFFFFF'); mf.cell(mx,1).fill=hdr
mf.merge_cells(start_row=mx,start_column=1,end_row=mx,end_column=1+len(GRADES)+1)
mxh=mx+1
mf.cell(mxh,1,"Prod\\Grd").font=F(size=9,bold=True,color='FFFFFF'); mf.cell(mxh,1).fill=sub; mf.cell(mxh,1).alignment=center; mf.cell(mxh,1).border=border
for j,g in enumerate(GRADES):
    c=mf.cell(mxh,2+j,g); c.font=F(bold=True,size=9,color=GCOLORS[g]); c.alignment=center; c.border=gborder(GCOLORS[g])
mf.cell(mxh,2+len(GRADES),"Total").font=F(size=9,bold=True,color='FFFFFF'); mf.cell(mxh,2+len(GRADES)).fill=sub; mf.cell(mxh,2+len(GRADES)).alignment=center; mf.cell(mxh,2+len(GRADES)).border=border
for i,p in enumerate(PRODUCTS):
    r=mxh+1+i
    c=mf.cell(r,1,p); c.font=F(bold=True,color='FFFFFF'); c.fill=PatternFill('solid',fgColor=PCOLORS[p]); c.alignment=center; c.border=border
    for j,g in enumerate(GRADES):
        cc=mf.cell(r,2+j,f'=SUMIFS(\'{P}\'!$G${LI_FIRST}:$G${LI_LAST},\'{P}\'!$B${LI_FIRST}:$B${LI_LAST},"{p}",\'{P}\'!$D${LI_FIRST}:$D${LI_LAST},"{g}")'); cc.alignment=center; cc.border=border
    rl=get_column_letter(2); rr=get_column_letter(1+len(GRADES))
    mf.cell(r,2+len(GRADES),f"=SUM({rl}{r}:{rr}{r})").font=F(bold=True); mf.cell(r,2+len(GRADES)).alignment=center; mf.cell(r,2+len(GRADES)).border=border
for j in range(len(GRADES)+1): mf.column_dimensions[get_column_letter(2+j)].width=7
mxtot=mxh+1+len(PRODUCTS)

# ===== PICK LIST: only the product/length/grade combos on this car =====
# Auto-generated from the layout. One line per inventory line item that is
# loaded; "Packs to Pick" = packs actually placed on the car for that combo.
LCOLORS={8:'3A6E9E',10:'2E7A5E',12:'7A5E2E',14:'6E3A7A',16:'8A3A3A',18:'2E7A6E',20:'4A6E2E'}
pl=mxtot+2
mf.cell(pl,1,"PICK LIST — packs to pull for this car (auto-generated from the layout)").font=F(size=10,bold=True,color='FFFFFF')
mf.cell(pl,1).fill=hdr; mf.merge_cells(start_row=pl,start_column=1,end_row=pl,end_column=5)
plh=pl+1
for j,h in enumerate(["#","Product","Length","Grade","Packs to Pick"]):
    c=mf.cell(plh,1+j,h); c.font=F(size=9,bold=True,color='FFFFFF'); c.fill=sub; c.alignment=center; c.border=border
plf=plh+1
nLI=LI_LAST-LI_FIRST+1
for i in range(nLI):
    r=plf+i; li=LI_FIRST+i
    # sequence number only when this line item is loaded
    mf.cell(r,1,f'=IF(OR(\'{P}\'!B{li}="",\'{P}\'!G{li}=0),"",\'{P}\'!A{li})').alignment=center; mf.cell(r,1).border=border
    mf.cell(r,2,f'=IF(OR(\'{P}\'!B{li}="",\'{P}\'!G{li}=0),"",\'{P}\'!B{li})').alignment=center; mf.cell(r,2).border=border
    mf.cell(r,3,f'=IF(OR(\'{P}\'!B{li}="",\'{P}\'!G{li}=0),"",\'{P}\'!C{li})').alignment=center; mf.cell(r,3).border=border
    mf.cell(r,4,f'=IF(OR(\'{P}\'!B{li}="",\'{P}\'!G{li}=0),"",\'{P}\'!D{li})').alignment=center; mf.cell(r,4).border=border
    mf.cell(r,5,f'=IF(OR(\'{P}\'!B{li}="",\'{P}\'!G{li}=0),"",\'{P}\'!G{li})').font=F(bold=True); mf.cell(r,5).alignment=center; mf.cell(r,5).border=border
pll=plf+nLI-1
ptot=pll+1
mf.cell(ptot,1,"TOTAL").font=F(bold=True); mf.cell(ptot,1).alignment=center; mf.cell(ptot,1).border=border
mf.merge_cells(start_row=ptot,start_column=1,end_row=ptot,end_column=4)
mf.cell(ptot,5,f"=SUM(E{plf}:E{pll})").font=F(bold=True); mf.cell(ptot,5).alignment=center; mf.cell(ptot,5).border=border
# color-code: product fill, length fill, grade border (keyed on each cell's own value)
PR=f"$B${plf}:$B${pll}"; LR=f"$C${plf}:$C${pll}"; GRr=f"$D${plf}:$D${pll}"
# Length = primary (fill); Product = secondary (thick border); Grade = tertiary (font color)
for L,hexc in LC.items():
    mf.conditional_formatting.add(LR, FormulaRule(formula=[f'C{plf}&""="{L}"'],fill=PatternFill('solid',fgColor=hexc),font=Font(name='Arial',bold=True,color='FFFFFF')))
for p,hexc in PCOLORS.items():
    mf.conditional_formatting.add(PR, FormulaRule(formula=[f'B{plf}="{p}"'],border=pborder(hexc),font=Font(name='Arial',bold=True,color='333333')))
for g,hexc in GCOLORS.items():
    mf.conditional_formatting.add(GRr, FormulaRule(formula=[f'D{plf}&""="{g}"'],font=Font(name='Arial',bold=True,color=hexc)))

# authoritative manifest column widths (last assignment wins across all tables)
mf.column_dimensions['A'].width=9
for col in 'BCDEFGHIJ': mf.column_dimensions[col].width=11
mf.column_dimensions['E'].width=13
mf.column_dimensions['K'].width=7
mf.column_dimensions['L'].width=13

# ---- print setup: fit on ONE 8.5x11 (Letter) page ----
from openpyxl.worksheet.properties import PageSetupProperties
mf.page_setup.orientation = "portrait"
mf.page_setup.paperSize = 1            # 1 = Letter (8.5 x 11)
mf.page_setup.fitToWidth = 1
mf.page_setup.fitToHeight = 1
mf.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
mf.page_margins.left = 0.2; mf.page_margins.right = 0.2
mf.page_margins.top = 0.25;  mf.page_margins.bottom = 0.25
mf.page_margins.header = 0.1; mf.page_margins.footer = 0.1
mf.print_options.horizontalCentered = True
mf.print_area = f"A1:L{ptot}"

# ===== PATTERN LIBRARY =====
LCOLORS={8:'3A6E9E',10:'2E7A5E',12:'7A5E2E',14:'6E3A7A',16:'8A3A3A',18:'2E7A6E',20:'4A6E2E'}
pt=wb.create_sheet("Pattern Library"); pt.sheet_view.showGridLines=False
pt['A1']="VALID 72-FT LENGTH PATTERNS"; pt['A1'].font=F(size=14,bold=True,color='2E5E20'); pt.merge_cells('A1:K1')
pt['A2']="Every length combination summing to 72 ft, color-coded by length. The 72-ft fit depends only on lengths; product & grade ride along on each pack."
pt['A2'].font=F(size=9,italic=True,color='666666'); pt.merge_cells('A2:K2')
pats=sv3.all_patterns(LENGTHS); pats.sort(key=lambda p:(len(p),[-x for x in p]))
ph=4
# header: # | Lengths | one column per length (colored) | Total
pt.cell(ph,1,"#").font=F(size=9,bold=True,color='FFFFFF'); pt.cell(ph,1).fill=sub; pt.cell(ph,1).alignment=center; pt.cell(ph,1).border=border
pt.cell(ph,2,"Lengths (ft)").font=F(size=9,bold=True,color='FFFFFF'); pt.cell(ph,2).fill=sub; pt.cell(ph,2).alignment=center; pt.cell(ph,2).border=border
for j,L in enumerate(LENGTHS):
    c=pt.cell(ph,3+j,f"{L}'"); c.font=F(size=9,bold=True,color='FFFFFF'); c.fill=PatternFill('solid',fgColor=LCOLORS[L]); c.alignment=center; c.border=border
tc=pt.cell(ph,3+len(LENGTHS),"Total"); tc.font=F(size=9,bold=True,color='FFFFFF'); tc.fill=sub; tc.alignment=center; tc.border=border
for idx,p in enumerate(pats):
    r=ph+1+idx
    pt.cell(r,1,idx+1).alignment=center; pt.cell(r,1).border=border
    # colored chips for the Lengths summary cell would be one fill only; keep text, color the count cells instead
    pt.cell(r,2," ".join(map(str,p))).alignment=left; pt.cell(r,2).border=border
    cnt={}
    for x in p: cnt[x]=cnt.get(x,0)+1
    for j,L in enumerate(LENGTHS):
        v=cnt.get(L,0); cc=pt.cell(r,3+j,v if v else None); cc.alignment=center; cc.border=border
        if v:  # color the cell by its length
            cc.fill=PatternFill('solid',fgColor=LCOLORS[L]); cc.font=F(bold=True,color='FFFFFF')
    pt.cell(r,3+len(LENGTHS),sum(p)).font=F(bold=True); pt.cell(r,3+len(LENGTHS)).alignment=center; pt.cell(r,3+len(LENGTHS)).border=border
# length color legend
lgp=ph+1+len(pats)+1
pt.cell(lgp,2,"Length colors:").font=F(size=9,bold=True,color='666666')
for j,L in enumerate(LENGTHS):
    c=pt.cell(lgp,3+j,f"{L}'"); c.font=F(bold=True,size=9,color='FFFFFF'); c.fill=PatternFill('solid',fgColor=LCOLORS[L]); c.alignment=center; c.border=border
pt.column_dimensions['A'].width=5; pt.column_dimensions['B'].width=24
for j in range(len(LENGTHS)): pt.column_dimensions[get_column_letter(3+j)].width=5
pt.column_dimensions[get_column_letter(3+len(LENGTHS))].width=7

# ===== VBA =====
v=wb.create_sheet("Auto-Solve (VBA)"); v.sheet_view.showGridLines=False
v['A1']="OPTIONAL: ONE-CLICK AUTO-SOLVER (VBA MACRO)"; v['A1'].font=F(size=14,bold=True,color='2E5E20'); v.merge_cells('A1:H1')
instr=["",
 "The Planner works WITHOUT macros (a solved example is loaded). For one-click solving:",
 "  1.  Save As ▸ Excel Macro-Enabled Workbook (*.xlsm).",
 "  2.  Alt+F11 ▸ File ▸ Import File ▸ CenterbeamSolver.bas  (or Insert ▸ Module and paste the code below).",
 "  3.  Developer ▸ Insert ▸ Button (Form Control); assign 'SolveLayout'. Add a 2nd button ▸ 'ClearAll'.",
 "  4.  SOLVE reads the line-item inventory, fills every selected row to 72 ft, and assigns product+grade",
 "      grouped so like packs stack in columns.  CLEAR resets the grid + line items. ClearGrid clears grid only.",
 "",
 "  Cell fill = product, border = grade (see legends on the Planner & Manifest).",
 "",
 "──────────────────────  VBA CODE — copy everything below  ──────────────────────"]
for i,t in enumerate(instr):
    c=v.cell(3+i,1,t); c.font=F(size=10,bold=("VBA CODE" in t)); v.merge_cells(start_row=3+i,start_column=1,end_row=3+i,end_column=8)

VBA=r'''Option Explicit
' === Centerbeam 72-ft Auto-Solver (product + length + grade) ===
Private Patterns As Collection
Private Found As Boolean
Private Chosen As Collection
Private RemCount() As Long
Private TypeLen() As Long
Private NumTypes As Long

Sub SolveLayout()
    Dim ws As Worksheet: Set ws = ThisWorkbook.Sheets("Planner")
    Dim prods As Variant: prods = Array("2x4", "2x6", "2x8", "2x10", "4x4", "4x6", "6x6")
    Dim grds As Variant:  grds = Array("1", "2", "3", "4", "2P", "MSR")
    Dim lens As Variant:  lens = Array(8, 10, 12, 14, 16, 18, 20)
    Const LIFIRST As Long = __LIFIRST__
    Const LILAST As Long = __LILAST__
    Const GRIDFIRST As Long = __GF__
    Const SLOTS As Long = 9
    Dim numRows As Long: numRows = ws.Range("C5").Value
    Dim target As Long: target = ws.Range("C6").Value

    ' --- read line items ---
    Dim nItems As Long: nItems = 0
    Dim itP() As String, itL() As Long, itG() As String, itQ() As Long
    ReDim itP(1 To LILAST - LIFIRST + 1)
    ReDim itL(1 To LILAST - LIFIRST + 1)
    ReDim itG(1 To LILAST - LIFIRST + 1)
    ReDim itQ(1 To LILAST - LIFIRST + 1)
    Dim r As Long
    For r = LIFIRST To LILAST
        Dim pp As String: pp = Trim(CStr(ws.Cells(r, 2).Value))
        Dim ll As Variant: ll = ws.Cells(r, 3).Value
        Dim gg As String: gg = Trim(CStr(ws.Cells(r, 4).Value))
        Dim qq As Variant: qq = ws.Cells(r, 5).Value
        If pp <> "" And IsNumeric(ll) And IsNumeric(qq) Then
            If CLng(qq) > 0 Then
                nItems = nItems + 1
                itP(nItems) = pp: itL(nItems) = CLng(ll): itG(nItems) = gg: itQ(nItems) = CLng(qq)
            End If
        End If
    Next r
    If nItems = 0 Then MsgBox "Enter some line items first.": Exit Sub

    ' --- length totals ---
    Dim nl As Long: nl = UBound(lens) + 1
    Dim lenTot() As Long: ReDim lenTot(0 To nl - 1)
    Dim i As Long, li As Long, total As Long
    For i = 1 To nItems
        For li = 0 To nl - 1
            If lens(li) = itL(i) Then lenTot(li) = lenTot(li) + itQ(i): total = total + itQ(i) * itL(i): Exit For
        Next li
    Next i

    ' --- length types (descending) ---
    NumTypes = 0
    For li = 0 To nl - 1
        If lenTot(li) > 0 Then NumTypes = NumTypes + 1
    Next li
    ReDim TypeLen(0 To NumTypes - 1): ReDim RemCount(0 To NumTypes - 1)
    Dim k As Long: k = 0
    For li = nl - 1 To 0 Step -1
        If lenTot(li) > 0 Then TypeLen(k) = lens(li): RemCount(k) = lenTot(li): k = k + 1
    Next li

    ' --- solve length layout ---
    Set Patterns = New Collection
    EnumPatterns 0, target, ""
    SortPatterns
    Dim fillable As Long
    If numRows < Int(total / target) Then fillable = numRows Else fillable = Int(total / target)
    Set Chosen = New Collection
    Found = False
    Solve 0, fillable, target

    Dim s As Long
    For r = 0 To 13
        For s = 0 To SLOTS - 1: ws.Cells(GRIDFIRST + r, 2 + s).ClearContents: Next s
    Next r
    If Not Found Then MsgBox "No exact 72-ft solution exists for these lengths and row count.", vbExclamation: Exit Sub

    ' --- order rows for column stacking ---
    Dim nR As Long: nR = Chosen.Count
    Dim ordered() As String: ReDim ordered(1 To nR)
    For r = 1 To nR: ordered(r) = Chosen(r): Next r
    Dim oa As Long, ob As Long, otmp As String
    For oa = 1 To nR - 1
        For ob = oa + 1 To nR
            If CompareRows(ordered(ob), ordered(oa)) < 0 Then otmp = ordered(oa): ordered(oa) = ordered(ob): ordered(ob) = otmp
        Next ob
    Next oa

    ' --- build product+grade queue per length (grouped product order then grade order) ---
    Dim maxQ As Long: maxQ = 0
    For li = 0 To nl - 1
        If lenTot(li) > maxQ Then maxQ = lenTot(li)
    Next li
    If maxQ < 1 Then maxQ = 1
    Dim qP() As String, qG() As String, qpos() As Long
    ReDim qP(0 To nl - 1, 1 To maxQ): ReDim qG(0 To nl - 1, 1 To maxQ): ReDim qpos(0 To nl - 1)
    Dim pix As Long, gix As Long, pos As Long, it As Long
    For li = 0 To nl - 1
        pos = 0
        For pix = 0 To UBound(prods)
            For gix = 0 To UBound(grds)
                For it = 1 To nItems
                    If itL(it) = lens(li) And itP(it) = prods(pix) And itG(it) = grds(gix) Then
                        Dim c As Long
                        For c = 1 To itQ(it): pos = pos + 1: qP(li, pos) = prods(pix): qG(li, pos) = grds(gix): Next c
                    End If
                Next it
            Next gix
        Next pix
        qpos(li) = 0
    Next li

    ' --- write rows ---
    For r = 1 To nR
        Dim arr() As String: arr = Split(Trim(ordered(r)), " ")
        For s = 0 To UBound(arr)
            If Len(arr(s)) > 0 Then
                Dim slotLen As Long: slotLen = CLng(arr(s))
                Dim lidx As Long: lidx = -1
                For li = 0 To nl - 1
                    If lens(li) = slotLen Then lidx = li: Exit For
                Next li
                If lidx >= 0 Then
                    If qpos(lidx) < lenTot(lidx) Then
                        qpos(lidx) = qpos(lidx) + 1
                        ws.Cells(GRIDFIRST + r - 1, 2 + s).Value = qP(lidx, qpos(lidx)) & "-" & slotLen & "-" & qG(lidx, qpos(lidx))
                    End If
                End If
            End If
        Next s
    Next r
    MsgBox "Solved: " & nR & " rows at 72 ft. Product+grade grouped & column-aligned.", vbInformation
End Sub

Private Sub EnumPatterns(ti As Long, remn As Long, cur As String)
    If remn = 0 Then Patterns.Add cur: Exit Sub
    If ti >= NumTypes Then Exit Sub
    Dim t As Long: t = TypeLen(ti)
    Dim maxN As Long
    If RemCount(ti) < Int(remn / t) Then maxN = RemCount(ti) Else maxN = Int(remn / t)
    Dim n As Long, j As Long, addStr As String
    For n = maxN To 0 Step -1
        addStr = cur
        For j = 1 To n: addStr = addStr & t & " ": Next j
        EnumPatterns ti + 1, remn - t * n, addStr
    Next n
End Sub

Private Function PieceCount(patStr As String) As Long
    Dim arr() As String: arr = Split(Trim(patStr), " ")
    Dim s As Long, c As Long
    For s = 0 To UBound(arr)
        If Len(arr(s)) > 0 Then c = c + 1
    Next s
    PieceCount = c
End Function

Private Function DistinctCount(patStr As String) As Long
    Dim arr() As String: arr = Split(Trim(patStr), " ")
    Dim seen As String, s As Long, c As Long: seen = "|"
    For s = 0 To UBound(arr)
        If Len(arr(s)) > 0 Then
            If InStr(seen, "|" & arr(s) & "|") = 0 Then seen = seen & arr(s) & "|": c = c + 1
        End If
    Next s
    DistinctCount = c
End Function

Private Sub SortPatterns()
    Dim n As Long: n = Patterns.Count
    If n < 2 Then Exit Sub
    Dim arr() As String: ReDim arr(1 To n)
    Dim i As Long: For i = 1 To n: arr(i) = Patterns(i): Next i
    Dim a As Long, b As Long, tmp As String
    For a = 1 To n - 1
        For b = a + 1 To n
            If PatBetter(arr(b), arr(a)) Then tmp = arr(a): arr(a) = arr(b): arr(b) = tmp
        Next b
    Next a
    Set Patterns = New Collection
    For i = 1 To n: Patterns.Add arr(i): Next i
End Sub

Private Function PatBetter(x As String, y As String) As Boolean
    Dim dx As Long, dy As Long: dx = DistinctCount(x): dy = DistinctCount(y)
    If dx <> dy Then PatBetter = (dx < dy): Exit Function
    Dim px As Long, py As Long: px = PieceCount(x): py = PieceCount(y)
    If px <> py Then PatBetter = (px < py): Exit Function
    PatBetter = (CompareRows(x, y) < 0)
End Function

Private Function CompareRows(x As String, y As String) As Long
    Dim ax() As String: ax = Split(Trim(x), " ")
    Dim ay() As String: ay = Split(Trim(y), " ")
    Dim nx As Long: nx = UBound(ax): Dim ny As Long: ny = UBound(ay)
    Dim mm As Long: If nx < ny Then mm = nx Else mm = ny
    Dim i As Long, vx As Long, vy As Long
    For i = 0 To mm
        vx = CLng(ax(i)): vy = CLng(ay(i))
        If vx <> vy Then CompareRows = IIf(vx > vy, -1, 1): Exit Function
    Next i
    If nx <> ny Then CompareRows = IIf(nx < ny, -1, 1) Else CompareRows = 0
End Function

Private Sub Solve(rowIdx As Long, fillable As Long, target As Long)
    If Found Then Exit Sub
    If rowIdx = fillable Then Found = True: Exit Sub
    Dim p As Variant, patStr As String
    For Each p In Patterns
        patStr = CStr(p)
        If CanUse(patStr) Then
            UsePat patStr, -1
            Chosen.Add patStr
            Solve rowIdx + 1, fillable, target
            If Found Then Exit Sub
            Chosen.Remove Chosen.Count
            UsePat patStr, 1
        End If
    Next p
End Sub

Private Function CanUse(patStr As String) As Boolean
    Dim need() As Long: ReDim need(0 To NumTypes - 1)
    Dim arr() As String: arr = Split(Trim(patStr), " ")
    Dim s As Long, i As Long, v As Long
    For s = 0 To UBound(arr)
        If Len(arr(s)) > 0 Then
            v = CLng(arr(s))
            For i = 0 To NumTypes - 1
                If TypeLen(i) = v Then need(i) = need(i) + 1: Exit For
            Next i
        End If
    Next s
    For i = 0 To NumTypes - 1
        If need(i) > RemCount(i) Then CanUse = False: Exit Function
    Next i
    CanUse = True
End Function

Private Sub UsePat(patStr As String, delta As Long)
    Dim arr() As String: arr = Split(Trim(patStr), " ")
    Dim s As Long, i As Long, v As Long
    For s = 0 To UBound(arr)
        If Len(arr(s)) > 0 Then
            v = CLng(arr(s))
            For i = 0 To NumTypes - 1
                If TypeLen(i) = v Then RemCount(i) = RemCount(i) + delta: Exit For
            Next i
        End If
    Next s
End Sub

Sub ClearGrid()
    Dim ws As Worksheet: Set ws = ThisWorkbook.Sheets("Planner")
    Dim r As Long, s As Long
    For r = 0 To 13
        For s = 0 To 8: ws.Cells(__GF__ + r, 2 + s).ClearContents: Next s
    Next r
End Sub

Sub ClearAll()
    Dim ws As Worksheet: Set ws = ThisWorkbook.Sheets("Planner")
    If MsgBox("Clear the whole layout and all line items?", vbQuestion + vbYesNo, "Start Over") <> vbYes Then Exit Sub
    Dim r As Long, s As Long
    For r = 0 To 13
        For s = 0 To 8: ws.Cells(__GF__ + r, 2 + s).ClearContents: Next s
    Next r
    For r = __LIFIRST__ To __LILAST__
        ws.Cells(r, 2).ClearContents: ws.Cells(r, 3).ClearContents
        ws.Cells(r, 4).ClearContents: ws.Cells(r, 5).ClearContents
    Next r
    MsgBox "Cleared. Enter new line items, then click Solve.", vbInformation
End Sub
'''
VBA=VBA.replace("__LIFIRST__",str(LI_FIRST)).replace("__LILAST__",str(LI_LAST)).replace("__GF__",str(GF))
vstart=3+len(instr)+1
for i,line in enumerate(VBA.split("\n")):
    c=v.cell(vstart+i,1,line if line else " "); c.font=Font(name='Consolas',size=9,color='1A1A1A'); c.fill=PatternFill('solid',fgColor='F2F2F2')
    v.merge_cells(start_row=vstart+i,start_column=1,end_row=vstart+i,end_column=8)
v.column_dimensions['A'].width=14
for col in ['B','C','D','E','F','G','H']: v.column_dimensions[col].width=12

wb.save('/home/claude/cb3.xlsx')
open('/mnt/user-data/outputs/CenterbeamSolver.bas','w').write('Attribute VB_Name = "CenterbeamSolver"\n'+VBA+"\n")
print("Manifest+Patterns+VBA added. Sheets:", wb.sheetnames)
