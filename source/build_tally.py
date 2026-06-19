#!/usr/bin/env python3
r"""
build_tally.py - Generate ../Centerbeam_Tally_Recommender.xlsm

Hybrid build:
  1. openpyxl writes ALL structure / formatting / data-validation / freeze-panes
     to a temporary .xlsx (fast, headless, no Excel required).
  2. A short Excel COM pass (pywin32) injects the VBA module + the sheet sort
     event + the two macro buttons, seeds an example run, and saves as .xlsm.

The VBA lives in source/tally_recommender.bas and is loaded via AddFromString,
so the macro logic stays readable and version-controlled.

Requirements:
  - openpyxl, pywin32  (pip install openpyxl pywin32)
  - Microsoft Excel installed, with Trust Center > "Trust access to the VBA
    project object model" enabled (one-time; already on for this project).

Run (use the REAL interpreter -- on Windows a bare `python` may resolve to the
Microsoft Store stub; use the full path to the python.org install):
  "%LOCALAPPDATA%\Programs\Python\Python313\python.exe" source\build_tally.py

Note: the COM pass uses DispatchEx (its own isolated Excel instance), so a
running Excel won't be hijacked -- but the output .xlsm must be CLOSED in every
Excel window or the final SaveAs fails with a file lock.
"""

import os
import sys
import tempfile

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation

# ----------------------------------------------------------------------------
# Paths
# ----------------------------------------------------------------------------
HERE = os.path.dirname(os.path.abspath(__file__))
REPO = os.path.dirname(HERE)
OUT = os.path.join(REPO, "Centerbeam_Tally_Recommender.xlsm")
BAS = os.path.join(HERE, "tally_recommender.bas")

# ----------------------------------------------------------------------------
# Palette (RGB hex - openpyxl uses normal RRGGBB, unlike COM's BGR integers)
# ----------------------------------------------------------------------------
NAVY, INK, WHITE = "1F4E79", "33414F", "FFFFFF"
HDRLT, STATUSF, NOTEF = "DDE6ED", "EAF0E6", "F0F4F8"
LC = ["BBD4EA", "C2E5C9", "FBE7B2", "E5CDEE", "FAC4BC", "BFEAE0", "DCE7AE"]  # 8..20 fills
LEN_NUMS = [8, 10, 12, 14, 16, 18, 20]
COLS = list("ABCDEFGHIJKL")
LEN_COLS = list("CDEFGHI")  # the 8'..20' columns

LEFT, CENTER, RIGHT = "left", "center", "right"


def bgr(hex_):
    """RRGGBB hex -> the BGR integer Excel COM expects for .RGB / .Color."""
    r, g, b = int(hex_[0:2], 16), int(hex_[2:4], 16), int(hex_[4:6], 16)
    return r + (g << 8) + (b << 16)


def font(bold=False, italic=False, size=11, color=INK):
    return Font(name="Calibri", size=size, bold=bold, italic=italic, color=color)


def solid(hex_):
    return PatternFill(fill_type="solid", fgColor=hex_)


def put(ws, addr, value=None, *, bold=False, italic=False, size=11, color=INK,
        fill=None, align=None, valign=None, wrap=False):
    c = ws[addr]
    if value is not None:
        c.value = value
    c.font = font(bold, italic, size, color)
    if fill:
        c.fill = solid(fill)
    if align or valign or wrap:
        c.alignment = Alignment(horizontal=align, vertical=valign, wrap_text=wrap)
    return c


def banner(ws, rng, value, *, fill=None, color=INK, bold=True, size=11,
           italic=False, align=LEFT):
    """Style every cell of a range, merge it, and set the top-left value."""
    f = font(bold, italic, size, color)
    pf = solid(fill) if fill else None
    for row in ws[rng]:
        for cc in row:
            cc.font = f
            if pf:
                cc.fill = pf
    ws.merge_cells(rng)
    top = rng.split(":")[0]
    ws[top] = value
    ws[top].alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)


def list_validation(ws, addr, items):
    dv = DataValidation(type="list", formula1='"%s"' % items, allow_blank=True,
                        showErrorMessage=False)
    ws.add_data_validation(dv)
    dv.add(addr)


def setup_grid(ws):
    widths = {"A": 8, "B": 42, "C": 6, "D": 6, "E": 6, "F": 6, "G": 6, "H": 6,
              "I": 6, "J": 10, "K": 9, "L": 13}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def color_length_headers(ws, row):
    for i, col in enumerate(LEN_COLS):
        ws["%s%d" % (col, row)].fill = solid(LC[i])


def fmt_block(ws, rng, numfmt=None):
    """Center a block (and optionally apply an integer number format) so the
    count/total columns read consistently across the sheet."""
    for row in ws[rng]:
        for c in row:
            c.alignment = Alignment(horizontal=CENTER, vertical="center")
            if numfmt:
                c.number_format = numfmt


# ----------------------------------------------------------------------------
# Stage 1: openpyxl builds the structure
# ----------------------------------------------------------------------------
def build_layout(path):
    wb = Workbook()

    # ---- Recommender ----
    rec = wb.active
    rec.title = "Recommender"
    setup_grid(rec)

    banner(rec, "A1:L1", "Centerbeam 72-ft Car  -  Lumber Tally Recommender",
           fill=NAVY, color=WHITE, bold=True, size=16, align=CENTER)
    rec.row_dimensions[1].height = 30
    banner(rec, "A2:L2",
           "Pick a car layout and match mode, check the lengths you want, then click Recommend Tallies. A mixed 7+5 car tallies each side separately.",
           italic=True, size=10, align=LEFT)

    banner(rec, "A4:B4", "SETTINGS", fill=HDRLT, bold=True)
    put(rec, "A5", "Car layout:", bold=True, align=RIGHT)
    put(rec, "B5", "720 ft - 5+5 (10 rows)")
    put(rec, "A6", "Match mode:", bold=True, align=RIGHT)
    put(rec, "B6", "Each selected must appear")   # default match mode (mode 2)
    banner(rec, "A7:L7",
           "Palette = use only these lengths   /   Each must appear = every picked length shows up in the car   /   + fillers = may add other lengths to finish a 72-ft row",
           italic=True, size=9)

    list_validation(rec, "B5", "720 ft - 5+5 (10 rows),864 ft - 7+5 mixed (12 rows),1008 ft - 7+7 (14 rows)")
    list_validation(rec, "B6", "Palette - use only selected,Each selected must appear,Each must appear + fillers")

    banner(rec, "A9:C9", "LENGTH PALETTE", fill=HDRLT, bold=True)
    put(rec, "A10", "Length", bold=True, align=CENTER)
    # Reactive caption (the VBA rewrites B10 to match the car layout: one column for a
    # symmetric car, "7-row / 5-row side" for a mixed 7+5 car).
    banner(rec, "B10:C10", "Check each length to load in this car",
           bold=True, size=9)
    for i, n in enumerate(LEN_NUMS):
        row = 11 + i
        put(rec, "A%d" % row, n, bold=True, fill=LC[i], align=CENTER)
        # Two Form Control checkboxes (added in the COM stage) link to B (7-row /
        # symmetric) and C (5-row) as TRUE/FALSE; ';;;' hides that text.
        c7 = put(rec, "B%d" % row, None, align=LEFT)
        c7.number_format = ";;;"
        c5 = put(rec, "C%d" % row, None, align=LEFT)
        c5.number_format = ";;;"

    put(rec, "A19", "Status:", bold=True)
    banner(rec, "B19:L19", "", fill=STATUSF)
    rec["B19"].alignment = Alignment(horizontal=LEFT, vertical="center", wrap_text=True)
    rec.row_dimensions[19].height = 28

    rec_hdr = ["#", "Tally (how to load the car)", "8'", "10'", "12'", "14'", "16'",
               "18'", "20'", "Total pcs", "Total ft", "OK?"]
    # Table A = the full car (symmetric) OR the 7-row side of a mixed car.
    # Table B = the 5-row side of a mixed car (empty for symmetric). Both banner
    # texts are set by the VBA at runtime to match the chosen car layout.
    banner(rec, "A21:L21", "RECOMMENDED FULL-CAR TALLIES   (click a length header 8'-20' to sort)",
           fill=NAVY, color=WHITE, bold=True)
    for c, txt in zip(COLS, rec_hdr):
        put(rec, "%s22" % c, txt, bold=True, fill=HDRLT)
    color_length_headers(rec, 22)
    banner(rec, "A35:L35", "5-ROW SIDE TALLIES", fill=NAVY, color=WHITE, bold=True)
    for c, txt in zip(COLS, rec_hdr):
        put(rec, "%s36" % c, txt, bold=True, fill=HDRLT)
    color_length_headers(rec, 36)

    banner(rec, "A49:L49",
           "Need the building blocks or to hand-build a custom blend?   ->   see the 'Row Patterns' tab.",
           fill=NOTEF, italic=True)

    # A faint, tongue-in-cheek copyright tucked at the bottom of the operating area.
    put(rec, "A51", "© 2026 Ken Paine  ·  all 72 feet reserved", italic=True, size=8, color="C8C8C8")

    # ---- Row Patterns ----
    pat = wb.create_sheet("Row Patterns")
    setup_grid(pat)
    banner(pat, "A1:J1",
           "VALID 72-FT ROW PATTERNS  -  building blocks  (set 'Rows to use' to hand-build; click a length header 8'-20' to sort)",
           fill=NAVY, color=WHITE, bold=True)
    pat.row_dimensions[1].height = 22
    pat_hdr = ["#", "Row = 72 ft", "8'", "10'", "12'", "14'", "16'", "18'", "20'",
               "Rows to use"]
    for c, txt in zip(COLS, pat_hdr):
        put(pat, "%s2" % c, txt, bold=True, fill=HDRLT)
    color_length_headers(pat, 2)
    pat.column_dimensions["J"].width = 12   # 'Rows to use' moved here (Pcs/Ft removed)
    pat.freeze_panes = "A3"   # <-- freeze the two header rows (one clean line)

    # ---- How to Use ----
    how = wb.create_sheet("How to Use")
    how.column_dimensions["A"].width = 104
    lines = [
        "CENTERBEAM 72-FT CAR  -  LUMBER TALLY RECOMMENDER",
        "",
        "WHAT IT DOES",
        "A 72-ft centerbeam car is loaded in rows; every row must be filled end-to-end to exactly 72 ft.",
        "This tool finds the length combinations that fill a 72-ft row, then recommends complete car tallies",
        "(720 ft = 10 rows, 864 ft = a mixed 7+5 car, or 1008 ft = 14 rows) using the lengths you choose.",
        "",
        "THE TABS",
        "Recommender  - set your options here and read the recommended full-car tallies.",
        "Row Patterns - every way to fill a single 72-ft row (building blocks); hand-build a custom car here.",
        "How to Use   - this sheet.",
        "",
        "HOW TO USE (Recommender tab)",
        "1.  Car layout: 5+5 (720 ft / 10 rows), 7+5 mixed (864 ft / 12 rows), or 7+7 (1008 ft / 14 rows).",
        "2.  Match mode:",
        "       - Palette - use only selected: tallies are built only from the lengths you turn on.",
        "       - Each selected must appear: every length you turn on must show up somewhere in the car.",
        "       - Each must appear + fillers: every selected length appears, other lengths may finish a row.",
        "3.  Length palette (reactive): a symmetric 5+5 / 7+7 car shows ONE column - just check each length to use.",
        "       Switch Car layout to 7+5 and a second column appears: LEFT box = 7-row side, RIGHT box = 5-row side,",
        "       and the two sides are tallied independently (504 ft + 360 ft).",
        "4.  Click Recommend Tallies.",
        "",
        "RECOMMENDED FULL-CAR TALLIES (Recommender tab)",
        "Ready-to-load cars. Each row shows the piece count for every length, the total pieces, the total feet",
        '(always rows x 72), and an OK check. "All 10 rows: 2x16, 2x20" means load every row the same way:',
        "two 16s + two 20s = 20 pcs 16 ft + 20 pcs 20 ft = 720 ft.",
        "",
        "ROW PATTERNS tab  (hand-build grid)",
        "Every way to fill a single 72-ft row from your lengths. To hand-build a car, type how many rows of",
        "each pattern in the Rows to use column; the YOUR HAND-BUILT TALLY line (directly below the last",
        "pattern) totals the pieces per length and shows the running rows / target (an OK appears when they",
        "match). Click a length column header (8 ft ... 20 ft) to sort a grid by that length, most first.",
        "This tab is reactive too: a symmetric car shows ONE grid (aim for 10 or 14 rows); a mixed 7+5 car",
        "shows TWO grids - the 7-row side (aim for 7 rows) and the 5-row side (aim for 5 rows) - built",
        "independently, exactly like the two recommendation tables.",
        "",
        "NOTES",
        "-  Only length controls the 72-ft fit; product and grade are not part of this tool.",
        "-  Click Enable Content when you open the file so the buttons, checkboxes and sorting work.",
        "-  Up to 15 recommendations and up to 101 row patterns are listed.",
    ]
    section_heads = {"WHAT IT DOES", "THE TABS", "HOW TO USE (Recommender tab)",
                     "RECOMMENDED FULL-CAR TALLIES (Recommender tab)",
                     "ROW PATTERNS tab  (hand-build grid)", "NOTES"}
    for i, text in enumerate(lines):
        if text:
            put(how, "A%d" % (i + 1), text,
                bold=text in section_heads, color=NAVY if text in section_heads else INK)
    put(how, "A1", lines[0], bold=True, size=14, color=NAVY)

    # ---- consolidate numeric formatting across the data blocks ----
    # (count + total columns centered as integers; runtime VBA writes preserve it)
    fmt_block(rec, "A23:A33", "0")
    fmt_block(rec, "A37:A47", "0")
    fmt_block(rec, "C23:K33", "0")
    fmt_block(rec, "C37:K47", "0")
    fmt_block(rec, "L23:L33")            # OK? is text, just center it
    fmt_block(rec, "L37:L47")
    # cover both stacked hand-build grids (the VBA can place a second grid well
    # below row 103 on a mixed car); keep counts centered as integers.
    fmt_block(pat, "A3:A240", "0")
    fmt_block(pat, "C3:J240", "0")       # lengths (C-I) + Rows to use (J)

    wb.save(path)


# ----------------------------------------------------------------------------
# Stage 2: Excel COM injects VBA + buttons, seeds, saves as .xlsm
# ----------------------------------------------------------------------------
# Click any hand-build grid's length header (8'-20', cols C-I) to sort it. The
# grids float at runtime-dependent rows, so detect a header by its col-B label
# ("Row = 72 ft") rather than a fixed row.
SHEET_EVENT = "\r\n".join([
    "Private Sub Worksheet_SelectionChange(ByVal Target As Range)",
    "    If Target.Cells.Count <> 1 Then Exit Sub",
    "    If Target.Column < 3 Or Target.Column > 9 Then Exit Sub",
    '    If CStr(Me.Cells(Target.Row, 2).Value) = "Row = 72 ft" Then',
    "        SortPatternsBy Target.Column, Target.Row",
    "    End If",
    "End Sub",
])

# Recommender tab: (1) click-to-sort the recommendation tables - table A header is
# row 22, table B (5-row side of a mixed car) header is row 36; length cols C-I.
# (2) Worksheet_Change makes the palette / tables / hand-build grids REACTIVE: any
# edit to the Car layout (B5) or Match mode (B6) re-runs the recommender.
REC_EVENT = "\r\n".join([
    "Private Sub Worksheet_SelectionChange(ByVal Target As Range)",
    "    If Target.Cells.Count <> 1 Then Exit Sub",
    "    If Target.Column < 3 Or Target.Column > 9 Then Exit Sub",
    "    If Target.Row = 22 Then SortRecsByRange Target.Column, 23, 33",
    "    If Target.Row = 36 Then SortRecsByRange Target.Column, 37, 47",
    "End Sub",
    "Private Sub Worksheet_Change(ByVal Target As Range)",
    "    If Intersect(Target, Me.Range(\"B5:B6\")) Is Nothing Then Exit Sub",
    "    Application.EnableEvents = False",
    "    RecommendTallies",
    "    Application.EnableEvents = True",
    "End Sub",
])

XL_MACRO_ENABLED = 52  # xlOpenXMLWorkbookMacroEnabled


def inject_macros(src_xlsx, out_xlsm):
    import win32com.client as win32

    with open(BAS, "r", encoding="utf-8") as fh:
        module_code = "\r\n".join(fh.read().splitlines())

    # DispatchEx forces a NEW, isolated Excel instance instead of attaching to one
    # the user already has open (plain Dispatch would, and our Quit() would then
    # close their session). The output .xlsm must still be closed everywhere or
    # SaveAs hits a file lock.
    excel = win32.DispatchEx("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    try:
        excel.AutomationSecurity = 1  # msoAutomationSecurityLow (allow macro Run)
    except Exception:
        pass

    wb = None
    try:
        print("      .. workbooks open before:", excel.Workbooks.Count, flush=True)
        print("      .. opening temp xlsx", flush=True)
        wb = excel.Workbooks.Open(src_xlsx)
        print("      .. opened; adding standard module", flush=True)

        # standard module
        comp = wb.VBProject.VBComponents.Add(1)  # vbext_ct_StdModule
        comp.Name = "TallyRecommender"
        comp.CodeModule.AddFromString(module_code)
        print("      .. module added; adding sheet event", flush=True)

        # sheet click-to-sort event on the Row Patterns sheet
        ws2 = wb.Worksheets("Row Patterns")
        wb.VBProject.VBComponents(ws2.CodeName).CodeModule.AddFromString(SHEET_EVENT)
        # same click-to-sort on the Recommender tab's recommendations table (row 22)
        rec_ws = wb.Worksheets("Recommender")
        wb.VBProject.VBComponents(rec_ws.CodeName).CodeModule.AddFromString(REC_EVENT)
        print("      .. events added; adding buttons", flush=True)

        # macro buttons on the Recommender sheet. Form Control buttons can't be
        # colored, so use rounded-rectangle shapes (same click->macro behavior via
        # OnAction) styled to the workbook scheme and laid out side by side.
        rec = wb.Worksheets("Recommender")
        MSO_ROUNDED_RECT = 5
        XL_CENTER = -4108  # xlHAlign/ xlVAlign Center

        def make_button(caption, macro, left, top, width, height,
                        fill_hex, text_hex, border_hex=None):
            shp = rec.Shapes.AddShape(MSO_ROUNDED_RECT, left, top, width, height)
            shp.Fill.Solid()
            shp.Fill.ForeColor.RGB = bgr(fill_hex)
            if border_hex:
                shp.Line.ForeColor.RGB = bgr(border_hex)
                shp.Line.Weight = 1.25
            else:
                shp.Line.Visible = False
            tf = shp.TextFrame
            tf.Characters().Text = caption
            tf.Characters().Font.Bold = True
            tf.Characters().Font.Size = 11
            tf.Characters().Font.Color = bgr(text_hex)
            tf.HorizontalAlignment = XL_CENTER
            tf.VerticalAlignment = XL_CENTER
            shp.OnAction = macro
            return shp

        top = rec.Range("D4").Top
        left = rec.Range("D4").Left
        make_button("Recommend Tallies", "RecommendTallies",
                    left, top, 150, 34, NAVY, WHITE)               # navy primary
        make_button("Clear", "ClearOutputsButton",
                    left + 160, top, 90, 34, HDRLT, INK, NAVY)     # light secondary
        print("      .. buttons added; adding palette checkboxes", flush=True)

        # Two length-palette checkboxes per length: "7-row" linked to col B
        # (also the symmetric palette) and "5-row" linked to col C, both within the
        # wide B column so they sit together. Linked cells use the ';;;' format from
        # the layout stage so only the boxes show next to the colored length chip.
        # Named cb7_<len> / cb5_<len> so the VBA can show/hide the 5-row column when
        # the car layout is symmetric (reactive palette).
        excel.EnableEvents = False  # keep the new Worksheet_Change quiet during build
        xl_on, xl_off = 1, -4146
        d7 = {16, 20}   # default 7-row / symmetric checks
        d5 = {16, 20}   # default 5-row checks
        for i, n in enumerate(LEN_NUMS):
            row = 11 + i
            cellB = rec.Range("B%d" % row)
            cb7 = rec.CheckBoxes().Add(cellB.Left + 4, cellB.Top + 1, 56, 15)
            cb7.Caption = "7-row"
            cb7.Name = "cb7_%d" % n
            cb7.LinkedCell = "$B$%d" % row
            cb7.Value = xl_on if n in d7 else xl_off
            rec.Range("B%d" % row).Value = (n in d7)
            cb5 = rec.CheckBoxes().Add(cellB.Left + 92, cellB.Top + 1, 56, 15)
            cb5.Caption = "5-row"
            cb5.Name = "cb5_%d" % n
            cb5.LinkedCell = "$C$%d" % row
            cb5.Value = xl_on if n in d5 else xl_off
            rec.Range("C%d" % row).Value = (n in d5)
        print("      .. checkboxes added; seeding (Run)", flush=True)

        # seed an example run so the file is populated on open
        rec.Activate()
        excel.Run("RecommendTallies")
        print("      .. seeded; saving .xlsm", flush=True)

        if os.path.exists(out_xlsm):
            os.remove(out_xlsm)
        wb.SaveAs(out_xlsm, XL_MACRO_ENABLED)
        has_vba = bool(wb.HasVBProject)
        print("      .. saved; HasVBProject =", has_vba, flush=True)
        wb.Close(False)
        return has_vba
    finally:
        try:
            if wb is not None and excel.Workbooks.Count:
                pass
        except Exception:
            pass
        excel.Quit()


def main():
    tmp = os.path.join(tempfile.gettempdir(), "tally_build_tmp.xlsx")
    print("[1/2] openpyxl: building layout ->", os.path.basename(tmp))
    build_layout(tmp)

    print("[2/2] Excel COM: injecting VBA + buttons, seeding, saving .xlsm")
    has_vba = inject_macros(tmp, OUT)

    try:
        os.remove(tmp)
    except OSError:
        pass

    size_kb = round(os.path.getsize(OUT) / 1024, 1)
    print("DONE ->", OUT)
    print("      HasVBProject:", has_vba, "| size:", size_kb, "KB")
    if not has_vba:
        print("WARNING: macros not present after save!", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
